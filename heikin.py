import pandas as pd
from openpyxl import load_workbook, Workbook
from pathlib import Path

# =========================
# ファイルパス設定
# =========================
desktop_path = Path.home() / "Desktop"

# ★ここが正しい（Desktop → Pythonフォルダ → xlsmファイル）
input_file = desktop_path / "Python" / "A組点数.xlsm"
input_file2 = desktop_path / "Python" / "B組点数.xlsm"
output_file = desktop_path / "平均点.xlsx"

# =========================
# Excelファイル読み込み
# =========================
wb = load_workbook(input_file, data_only=True, keep_vba=True)

# シート取得
sheet_a = wb["A組点数"]
sheet_b = wb["B組点数"]

# =========================
# 点数データ取得関数
# =========================
def get_scores(sheet):
    scores = []

    for row in sheet.iter_rows(
        min_row=3, max_row=8,
        min_col=2, max_col=6,
        values_only=True
    ):
        scores.append(list(row))

    df = pd.DataFrame(
        scores,
        columns=["数学", "国語", "社会", "英語", "理科"]
    )

    return df

# A組・B組データ取得
df_a = get_scores(sheet_a)
df_b = get_scores(sheet_b)

# =========================
# データ結合
# =========================
all_scores = pd.concat([df_a, df_b], ignore_index=True)

# =========================
# 平均計算
# =========================
average_scores = all_scores.mean()

# =========================
# 新しいExcel作成
# =========================
new_wb = Workbook()
ws = new_wb.active
ws.title = "平均点"

subjects = ["数学", "国語", "社会", "英語", "理科"]

# ヘッダー
for col_num, subject in enumerate(subjects, start=2):
    ws.cell(row=1, column=col_num, value=subject)

# 平均行
ws.cell(row=2, column=1, value="平均")

for col_num, subject in enumerate(subjects, start=2):
    ws.cell(
        row=2,
        column=col_num,
        value=round(average_scores[subject], 2)
    )

# =========================
# 保存
# =========================
new_wb.save(output_file)

print(f"平均点ファイルを作成しました: {output_file}")